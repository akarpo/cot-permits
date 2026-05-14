#!/usr/bin/env python3
"""
Scrape City of Troy, MI "Permits Issued" records into two timestamped Excel files:

  1. troy_permits_metro_act_<timestamp>.xlsx
       Right of Way permits whose Permit Type is "ROW, METRO ACT".
       Any cell mentioning "Fiber" is highlighted yellow.

  2. troy_permits_all_other_<timestamp>.xlsx
       Every other permit in the system (all types, all history), Metro Act excluded.

It talks directly to the JSON endpoint the site's own JavaScript uses
(`/PermitsIssued/Results`) -- no browser, no Selenium, no manual steps.

NOTE: file 2 is large (~266k rows / ~10,600 pages); a full run typically takes
2+ hours depending on --delay. Use --max-pages for a quick test run first.
File 1 is written as soon as its (~2 min) scrape finishes, so it is available
while file 2 keeps running. A Ctrl+C still writes whatever was collected.

Usage:
    python scrape_permits.py                       # full run
    python scrape_permits.py --max-pages 5         # quick test
    python scrape_permits.py --delay 0.5 --output-dir ~/Desktop
"""

import argparse
import os
import re
import sys
import time
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_URL = "https://apps.troymi.gov/PermitsIssued"
RESULTS_URL = f"{BASE_URL}/Results"

# The site returns 403 to requests that don't look like a real browser.
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "X-Requested-With": "XMLHttpRequest",
}

METRO_ACT_MARKER = "METRO ACT"   # matched (case-insensitive) in the Permit Type column
FIBER_MARKER = "FIBER"           # cells containing this are highlighted in file 1
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFFF00")  # yellow
BOLD = Font(bold=True)


def clean(text):
    """Collapse all whitespace so each spreadsheet cell is a tidy single line."""
    return re.sub(r"\s+", " ", text).strip()


def fetch_page(session, page_number, permit_type, max_retries=3):
    """Fetch one page of results, returning the raw table HTML fragment."""
    params = {"PageNumber": page_number, "PermitType": permit_type}
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(RESULTS_URL, params=params, timeout=30)
            resp.raise_for_status()
            payload = resp.json()
            if not payload.get("success", False):
                raise RuntimeError(f"server reported failure: {payload.get('message')!r}")
            return payload["data"]["table"]
        except (requests.RequestException, ValueError, KeyError, RuntimeError) as e:
            last_error = e
            wait = 2 * attempt
            print(f"  page {page_number}: attempt {attempt} failed ({e}); "
                  f"retrying in {wait}s...")
            time.sleep(wait)
    raise RuntimeError(f"page {page_number} failed after {max_retries} attempts: {last_error}")


def parse_table(table_html):
    """Parse a table fragment into (headers, list_of_row_lists)."""
    soup = BeautifulSoup(table_html, "lxml")
    table = soup.find("table")
    if table is None:
        return [], []
    headers, rows = [], []
    for tr in table.find_all("tr"):
        ths = tr.find_all("th")
        if ths:
            headers = [clean(th.get_text()) for th in ths]
            continue
        tds = tr.find_all("td")
        if tds:
            rows.append([clean(td.get_text()) for td in tds])
    return headers, rows


def get_total_pages(table_html):
    """Read the highest page number from the #Pagination block (1 if absent)."""
    pager = BeautifulSoup(table_html, "lxml").find(id="Pagination")
    if pager is None:
        return 1
    numbers = [int(n) for n in re.findall(r"\d+", pager.get_text(" "))]
    return max(numbers) if numbers else 1


def scrape(session, permit_type, label, keep_row, collector,
           delay, max_pages, progress_every):
    """
    Scrape every page for one PermitType filter.

    keep_row(row, permit_type_idx) -> bool decides which rows to retain.
    Retained rows land in collector["rows"] and headers in collector["headers"]
    as the scrape runs, so a Ctrl+C or crash mid-scrape still leaves partial
    data for the caller to save. Pages that fail every retry are skipped and
    recorded in collector["failed_pages"] rather than aborting the whole run.
    """
    first_html = fetch_page(session, 1, permit_type)
    headers, rows = parse_table(first_html)
    collector["headers"] = headers
    total_pages = get_total_pages(first_html)
    last_page = min(total_pages, max_pages) if max_pages else total_pages

    pt_idx = headers.index("Permit Type") if "Permit Type" in headers else 0
    seen = set()  # permit numbers already collected -- guards against repeats

    def take(page_rows):
        for row in page_rows:
            # Permit Number (col 1) is unique; fall back to the whole row.
            key = row[1] if len(row) > 1 else tuple(row)
            if key in seen:
                continue
            seen.add(key)
            if keep_row(row, pt_idx):
                collector["rows"].append(row)

    take(rows)
    start = time.time()
    print(f"[{label}] {total_pages} pages available; scraping {last_page}.")

    page = 2
    while page <= last_page:
        try:
            page_rows = parse_table(fetch_page(session, page, permit_type))[1]
        except RuntimeError as e:
            print(f"[{label}] SKIPPING page {page}: {e}")
            collector["failed_pages"].append(page)
            page += 1
            continue

        before = len(seen)
        take(page_rows)
        if len(seen) == before and page_rows:
            # The server clamps an out-of-range PageNumber to the last page,
            # so a page that adds nothing new means we've reached the end.
            print(f"[{label}] page {page} added no new rows -- stopping early.")
            break

        if page % progress_every == 0 or page == last_page:
            elapsed = time.time() - start
            rate = (page - 1) / elapsed if elapsed else 0
            eta = (last_page - page) / rate / 60 if rate else 0
            print(f"[{label}] page {page}/{last_page} | "
                  f"{len(collector['rows'])} rows kept | ETA {eta:.1f} min")
        page += 1
        time.sleep(delay)

    return collector


def column_width(header):
    """A reasonable fixed width for a column, given its header text."""
    if header == "Work Description":
        return 70
    return min(max(len(header) + 4, 14), 30)


def apply_widths(ws, headers):
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width(header)


def write_metro_act(path, headers, rows):
    """File 1 -- small dataset; highlight every cell that mentions Fiber."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ROW Metro Act"
    apply_widths(ws, headers)
    ws.freeze_panes = "A2"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD

    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            if cell.value and FIBER_MARKER in str(cell.value).upper():
                cell.fill = HIGHLIGHT_FILL

    wb.save(path)


def write_all_other(path, headers, rows):
    """File 2 -- large dataset; streamed via a write-only workbook."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("All Other Permits")
    apply_widths(ws, headers)
    ws.freeze_panes = "A2"

    header_cells = []
    for h in headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = BOLD
        header_cells.append(cell)
    ws.append(header_cells)

    for row in rows:
        ws.append(row)

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--output-dir", default=".",
                        help="directory to write the two .xlsx files into")
    parser.add_argument("--delay", type=float, default=0.25,
                        help="seconds to wait between page requests (default: 0.25)")
    parser.add_argument("--max-pages", type=int, default=None,
                        help="cap pages per scrape -- useful for a quick test run")
    parser.add_argument("--progress-every", type=int, default=50,
                        help="print a progress line every N pages (default: 50)")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    metro_path = os.path.join(args.output_dir, f"troy_permits_metro_act_{timestamp}.xlsx")
    other_path = os.path.join(args.output_dir, f"troy_permits_all_other_{timestamp}.xlsx")

    session = requests.Session()
    session.headers.update(HEADERS)

    def is_metro_act(row, idx):
        return len(row) > idx and METRO_ACT_MARKER in row[idx].upper()

    def not_metro_act(row, idx):
        return not is_metro_act(row, idx)

    metro = {"headers": [], "rows": [], "failed_pages": [], "written": False}
    other = {"headers": [], "rows": [], "failed_pages": []}

    try:
        print("=== File 1: Right of Way (Metro Act) ===")
        scrape(session, "Right of Way (ROW)", "METRO ACT", is_metro_act, metro,
               args.delay, args.max_pages, args.progress_every)
        print(f"  -> {len(metro['rows'])} Metro Act permits")
        # Write file 1 as soon as its scrape finishes so it lands on disk
        # while the long file 2 scrape keeps running.
        if metro["rows"]:
            write_metro_act(metro_path, metro["headers"], metro["rows"])
            metro["written"] = True
            print(f"Saved {len(metro['rows'])} rows -> {metro_path}\n")

        print("=== File 2: all other permits (this is the long one) ===")
        scrape(session, "All", "ALL OTHER", not_metro_act, other,
               args.delay, args.max_pages, args.progress_every)
        print(f"  -> {len(other['rows'])} non-Metro-Act permits\n")

    except KeyboardInterrupt:
        print("\nInterrupted -- writing whatever was collected so far...")
    except Exception as e:  # noqa: BLE001 - top-level guard so partial data is saved
        print(f"\nError: {e}", file=sys.stderr)
    finally:
        # File 1 is normally written above as soon as its scrape finishes;
        # this fallback only fires if the run was interrupted partway through
        # the file 1 scrape itself.
        if metro["rows"] and not metro["written"]:
            write_metro_act(metro_path, metro["headers"], metro["rows"])
            print(f"Saved {len(metro['rows'])} rows -> {metro_path}")
        if other["rows"]:
            write_all_other(other_path, other["headers"], other["rows"])
            print(f"Saved {len(other['rows'])} rows -> {other_path}")
        for label, c in (("Metro Act", metro), ("All other", other)):
            if c["failed_pages"]:
                print(f"WARNING [{label}]: {len(c['failed_pages'])} page(s) failed "
                      f"and were skipped: {c['failed_pages']}")
        if not metro["rows"] and not other["rows"]:
            print("Nothing collected; no files written.")


if __name__ == "__main__":
    main()
